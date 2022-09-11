import requests
import openpyxl as px
import cv2
import os
import subprocess
import datetime
import glob, re
import numpy as np
import unicodedata
import pprint
import time
import urllib.error
import urllib.request
import ffmpeg
from bs4 import BeautifulSoup
from yt_dlp import YoutubeDL
from yt_dlp.utils import DownloadError
from pytube import YouTube
from moviepy.editor import *
from moviepy.config import change_settings
change_settings({"IMAGEMAGICK_BINARY": r"C:\\ImageMagick\\magick.exe"})
from PIL import Image, ImageDraw, ImageFont
from pilmoji import Pilmoji
from openpyxl.styles import Font

#  ファイル作成コマンド定数
LIST_FILE_NAME='movielist.txt'
FILE_HEADER="file '"
FILE_FOOTER="'"
RETURN = '\n'
SEARCH_OBJ_SENTENCE='./download/*.*'
#  FFMPEGコマンド
FFMPEG_HEADER='ffmpeg -safe 0 -f concat -i '
FFMPEG_MIDDLE=' -c:v copy -c:a copy -map 0:v -map 0:a '
now = datetime.datetime.now()
today = datetime.date.today()
twoday=datetime.timedelta(days=2) 
downloadday=today-twoday
OUT_FILE_NAME= downloadday.strftime('%Y%m%d_')+'dailyranking.mp4'
#動画タイトル、url、再生数、再生時間、所属グループ、タイトルの長さ(バイト数)
titles = []
urls = []
views = []
times = []
groups = []
bytes = []
#所属グループ判定用配列
offices_niji = ['にじさんじ(統合後)','にじさんじ(SEEDs出身)','にじさんじ(1・2期生)','NIJISANJI EN','NIJISANJI ID','NIJISANJI KR']
offices_holo = ['hololive English','ホロライブ','hololive Indonesia']
#フォント
fontpass = './GenJyuuGothicL-Bold.ttf'

# 動画情報取得
def get_video_info():
  #VTuberランキング
  url = 'https://virtual-youtuber.userlocal.jp/movies?date='+str(downloadday)
  # url = 'https://virtual-youtuber.userlocal.jp/movies?date=2022-07-22'
  #GETリクエストを送信
  reqs = requests.get(url)
  #エクセルを取得
  wb = px.Workbook()
  ws = wb.active
  #URLをテキスト化し、解析を行う。その後BeautifulSoupオブジェクトを作る
  soup = BeautifulSoup(reqs.text, 'html.parser')
  #動画タイトル、url、再生数、再生時間、所属グループを取得する
  tag_list = soup.select('tr[data-title]')
  for tag in tag_list:
    title = tag.get('data-title')
    if "[LIVE]" in title:
      title = title[6:]
    url = tag.get('data-video-url')
    view = tag.find('div', class_="text-nowrap").get_text(strip=True, separator="|")
    try:
      group = tag.find('td', class_="col-info vertical p-3").find('img').get('alt')
    except:
      group = ''
      pass
    if (group in offices_niji or group in offices_holo) and ("#shorts" not in title) and ("【アニメ】" not in title):
      yt = YouTube(url)
      try:
        # 再生時間が5分未満はランク外とする
        if(yt.length > 300):
          titles.append(yt.title)
          urls.append(url)
          views.append(view.split('|')[0])
          times.append(yt.length)
          groups.append(group)
          bytes.append(get_str_width(yt.title))
      except:
        pass
  #エクセルに書き込む
  for i in range(len(titles)):
    ws.cell(row=i+1, column=1, value=titles[i])
    ws.cell(row=i+1, column=2, value=bytes[i])
    ws.cell(row=i+1, column=3, value=urls[i])
    ws.cell(row=i+1, column=4, value=views[i])
    ws.cell(row=i+1, column=5, value=times[i])
    ws.cell(row=i+1, column=6, value=groups[i])
    ws['A'+str(i+1)].font = Font(name='ＭＳ Ｐゴシック')
  wb.save("./result/"+downloadday.strftime('%Y%m%d_')+"result.xlsx")

# 動画をダウンロードして編集
def video_download(i, max, filenum):
  #上位20までの動画をダウンロードする
  while i < max:
    # try:
      # タイトル分割用変数
      arr=[]
      oneline=False
      twoline=False
      threeline=False
      byte_cnt=0
      result_str=''
      
      yt = YouTube(urls[i])
      # サムネイル画像を取得
      url='https://img.youtube.com/vi/'+yt.video_id+'/sddefault.jpg'
      png_savePath = './download/'+str(filenum).zfill(2)+'_kirinuki.png'
      download_thumbnail(url, png_savePath)
      # 頭出し1秒の動画を作成
      head_savePath = './download/'+str(filenum).zfill(2)+'_head.mp4'
      cmd = 'ffmpeg -loop 1 -i '+png_savePath+' -i dodon.mp3 -vcodec h264_nvenc -pix_fmt yuv420p -color_primaries bt709 -color_trc bt709 -colorspace bt709 -t 00:00:01.58 -r 60 -y '+head_savePath
      subprocess.call(cmd)
      # タイトルを1文字ずつチェックして指定のbyte数で分割する
      for char in yt.title:
        byte_cnt += get_char_width(char)
        result_str += char
        if get_str_width(yt.title) < 54 and oneline==False:
          arr.append(yt.title)
          oneline=True
        if byte_cnt > 54 and twoline==False:
          arr.append(result_str)
          result_str=''
          twoline=True
        if byte_cnt > 108 and threeline==False:
          arr.append(result_str)
          result_str=''
          threeline=True
        if byte_cnt == get_str_width(yt.title) and result_str not in arr:
          arr.append(result_str)
      # 空の要素を削除して無駄な改行をなくす
      arr = list(filter(None, arr))
      # タイトルと再生回数をテキストファイルに書き込み
      f = open('title.txt', 'w', encoding='UTF-8')
      g = open('views.txt', 'w', encoding='UTF-8')
      # タイトル位置調整
      if (i>8):
        w = 330
      else:
        w = 250
      # タイトルの長さによって位置調整および折り返し処理をする
      if((get_str_width(yt.title)>54 and get_str_width(yt.title)<108 and len(arr) == 2)
         or (get_str_width(yt.title)>109 and len(arr) == 2)):
        h = 30  # タイトル位置調整
        v_h = 170 # 再生回数位置調整
        yt.title = arr[0]+'\n'+arr[1] # タイトルを2行にする
        f.write(yt.title)
        f.close()
        convert_text_to_img(yt.title)
      elif(get_str_width(yt.title)>108 and get_str_width(yt.title)<162 and len(arr) == 3):
        h = 30  # タイトル位置調整
        v_h = 230 # 再生回数位置調整
        yt.title = arr[0]+'\n'+arr[1]+'\n'+arr[2] # タイトルを3行にする
        f.write(yt.title)
        f.close()
        convert_text_to_img(yt.title)
      else:
        h = 60
        v_h = 170 # 再生回数位置調整
        f.write(yt.title)
        f.close()
        convert_text_to_img(yt.title)
      g.write('再生回数：'+views[i]+'回')
      g.close()
      #切り抜き開始位置と終了位置を設定(1動画1分)
      startTime = yt.length/2
      endTime = startTime + 60
      #動画を切り抜いてダウンロード
      outputpath = './download/'+str(filenum).zfill(2)+'_clip.mp4'
      try:
        cmd = 'yt-dlp -o '+outputpath+' -f bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best '+urls[i]+' --download-sections *'+str(startTime)+'-'+str(endTime)
      except DownloadError:
        # os.remove(png_savePath)
        # os.remove(head_savePath)
        del urls[i]
        del views[i]
        video_download(i, 20, filenum)
      subprocess.call(cmd.split())
      #順位、タイトル、再生回数を動画に重ねて、フェードインとフェードアウト処理をする(60fpsで出力)
      tmp1Path = './download/'+str(filenum).zfill(2)+'_tmp1.mp4'
      tmp2Path = './download/'+str(filenum).zfill(2)+'_tmp2.mp4'
      savePath = './download/'+str(filenum).zfill(2)+'_kirinuki.mp4'
      cmd = 'ffmpeg -i '+outputpath+' -i title.png -filter_complex overlay='+str(w)+':'+str(h)+' -c:v h264_nvenc -y '+tmp1Path
      subprocess.call(cmd)
      cmd = 'ffmpeg -i '+tmp1Path+' -vf "[in]drawtext=fontfile='+fontpass+':x=30:y=30:fontsize=120:fontcolor=white:bordercolor=black:borderw=4:text='+str(i+1)+'位,  drawtext=fontfile='+fontpass+':x=30:y='+str(v_h)+':fontsize=50:fontcolor=white:bordercolor=black:borderw=4:textfile=./views.txt, fade=t=in:st=0:d=3,fade=t=out:st=57:d=3[out]" -c:a copy -c:v h264_nvenc -r 60 -y '+tmp2Path
      subprocess.call(cmd)
      # 頭出し1秒動画と結合
      clip1 = VideoFileClip(head_savePath)
      clip2 = VideoFileClip(tmp2Path)
      clip = concatenate_videoclips([clip1, clip2])
      clip.write_videofile(savePath,fps=60, threads=6, codec="h264_nvenc")
      #保存後、元動画を削除(後処理)
      os.remove(png_savePath)
      os.remove(head_savePath)
      os.remove(outputpath)
      os.remove(tmp1Path)
      os.remove(tmp2Path)
      i+=1
      filenum+=1
    # except:
    #   del urls[i]
    #   del views[i]
    #   pass

# ダウンロードした動画をマージ
def video_merge():
  # downloadフォルダ配下にある動画パスをテキストファイルに書き込む
  outfile = open(LIST_FILE_NAME, 'w')
  # downloadフォルダ以下の要素分だけループ(降順)
  for f in sorted(glob.glob(SEARCH_OBJ_SENTENCE, recursive=True), reverse=True):
      if re.search('.*.(mp4|MTS)', f):
          batCom = FILE_HEADER + str(f) + FILE_FOOTER + RETURN
          outfile.write(batCom)
  outfile.close()
  # FFMPEGコマンドの実行(複数の動画ファイルを1つにまとめる)
  # ffmpeg -safe 0 -f concat -i [定義ファイル名] -c:v copy -c:a copy -map 0:v -map 0:a [出力ファイル名]
  cmd = FFMPEG_HEADER + LIST_FILE_NAME + FFMPEG_MIDDLE + OUT_FILE_NAME
  subprocess.call(cmd.split())
  # kirinuki.mp4を削除(後処理)
  for f in sorted(glob.glob(SEARCH_OBJ_SENTENCE, recursive=True), reverse=True):
      if re.search('.*.(mp4|MTS)', f):
            os.remove(f)

# 動画のFPSをチェック(デバッグ用)
def check_fps():
    for f in sorted(glob.glob(SEARCH_OBJ_SENTENCE, recursive=True), reverse=True):
        if re.search('.*.(mp4|MTS)', f):
          cap = cv2.VideoCapture(f)
          fps = cap.get(cv2.CAP_PROP_FPS)
          print(fps)

# 動画タイトルをpng画像へ変換(絵文字対策)
def convert_text_to_img(text):
  # RGB, 画像サイズ, 背景色を設定
  im = Image.new("RGB", (1750, 180), (100, 100, 100,0))
  # PCローカルのフォントへのパスと、フォントサイズを指定
  font = ImageFont.truetype(fontpass, 50)
  # 文字の描画(描画位置、描画する文字、文字色、フォントを指定)
  with Pilmoji(im) as pilmoji:
    pilmoji.text((0, 0), text, (255, 255, 255), font=font, stroke_width=4, stroke_fill='black')
  # ファイルに出力
  im.save("./tmp.png")
  # 背景の透過処理
  path = "./tmp.png"
  src = cv2.imread(path)
  mask = np.all(src[:,:,:] == [100, 100, 100], axis=-1)
  dst = cv2.cvtColor(src, cv2.COLOR_BGR2BGRA)
  dst[mask,3] = 0
  cv2.imwrite("./title.png", dst)
  os.remove(path)

# 全角半角を区別してバイト数を取得
def get_char_width(c):
  data = unicodedata.east_asian_width(c)
  if data in ['Na', 'H']:
    return 1
  else:
    return 2
      
# 文字列長取得関数
def get_str_width(s):
  return sum([get_char_width(c) for c in s])

# 動画サムネイル取得
def download_thumbnail(url, dst_path):
  try:
    with urllib.request.urlopen(url) as web_file:
      data = web_file.read()
      with open(dst_path, mode='wb') as local_file:
          local_file.write(data)
      # 画像をリサイズ
      im = Image.open(dst_path)
      resized = im.resize((int(im.width*3), int(im.height*2.25)))
      resized.save(dst_path, quality=100)
  except urllib.error.URLError as e:
      print(e)
  
def test():
  for i in range(20):
    yt = YouTube(urls[i])
    s = yt.title.encode('utf-8')
    d = s.decode('utf-8')
    print(d)
    # # サムネイル画像を取得
    # url='https://img.youtube.com/vi/'+yt.video_id+'/sddefault.jpg'
    # png_savePath = './download/'+str(i+1).zfill(2)+'_kirinuki.png'
    # download_thumbnail(url, png_savePath)

#メイン処理------------------------------------------------
get_video_info()
video_download(0, 10, 1)
video_merge()
# check_fps()
# test()
